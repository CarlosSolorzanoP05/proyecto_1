from django.contrib import messages
from django.shortcuts import redirect
from django.http import HttpResponse
import io
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin, UserPassesTestMixin
from django.views.generic import CreateView, UpdateView
from django.shortcuts import redirect
from django.contrib import messages


# ─────────────────────────────────────────────
#  StaffOrAdminRequiredMixin
#  REQUERIMIENTO 3 - Permisos de Productos (Ver y Filtrar solamente
#  para el rol "Usuario"). Se implementa con LoginRequiredMixin +
#  UserPassesTestMixin (tal como pide el enunciado) en lugar de
#  ModulePermissionRequiredMixin, porque este último solo distingue
#  "tiene o no tiene el permiso view_products" (que Usuario SÍ tiene,
#  para poder listar/filtrar productos) y no alcanza para bloquear
#  selectivamente crear/editar/eliminar sin tocar ese permiso de
#  lectura compartido.
# ─────────────────────────────────────────────
class StaffOrAdminRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    """
    Permite el acceso solo a superusuarios o a usuarios que pertenezcan
    a los grupos/rangos 'Administrador' o 'Trabajador'.

    Uso (billing/views.py):
        class ProductCreateView(StaffOrAdminRequiredMixin, ModulePermissionRequiredMixin, ...):
            ...

    El rol "Usuario" (rango estándar autoregistrado en /signup/) NUNCA
    pasa este test, por lo que queda bloqueado de crear, editar o
    eliminar Productos, aunque sí pueda seguir viendo el listado y el
    detalle (esas vistas no usan este mixin).
    """
    allowed_groups = ('Administrador', 'Trabajador')
    permission_denied_message = 'No tienes permisos (rango Administrador/Trabajador requerido) para realizar esta acción.'

    def test_func(self):
        user = self.request.user
        if not user.is_authenticated:
            return False
        if user.is_superuser:
            return True
        return user.groups.filter(name__in=self.allowed_groups).exists()

    def handle_no_permission(self):
        if not self.request.user.is_authenticated:
            return super().handle_no_permission()
        messages.error(self.request, self.permission_denied_message)
        return redirect('billing:product_list')


# ─────────────────────────────────────────────
#  StaffRequiredMixin  (sin cambios)
# ─────────────────────────────────────────────
class StaffRequiredMixin:
    """
    Mixin que verifica si el usuario es miembro del staff.
    Si no es staff, redirige con mensaje de error.
    """
    staff_redirect_url = '/'
    staff_error_message = 'You do not have permission to perform this action. Staff access required.'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_staff:
            messages.error(request, self.staff_error_message)
            return redirect(self.staff_redirect_url)
        return super().dispatch(request, *args, **kwargs)


# ─────────────────────────────────────────────
#  Helpers de exportación genéricos (PDF / Excel)
#  Usados tanto por ExportMixin (CBV) como por las
#  vistas basadas en función (FBV) a través de
#  export_list_response().
# ─────────────────────────────────────────────

def resolve_field(obj, field):
    """Obtiene el valor de un campo simple, relación (__) o callable."""
    if callable(field):
        return field(obj)
    value = obj
    for part in field.split('__'):
        value = getattr(value, part, '')
        if callable(value):
            value = value()
    return value if value is not None else ''


def build_rows(qs, fields):
    """Construye las filas de datos (lista de listas de strings)."""
    rows = []
    for obj in qs:
        row = [str(resolve_field(obj, f)) for f in fields]
        rows.append(row)
    return rows


def build_excel_response(qs, filename, headers, fields):
    """Genera un HttpResponse .xlsx a partir de un queryset."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse(
            'openpyxl no está instalado. Ejecuta: pip install openpyxl',
            status=500,
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename[:31]   # max 31 chars en Excel

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(color='FFFFFF', bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(build_rows(qs, fields), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def build_pdf_response(qs, filename, headers, fields):
    """Genera un HttpResponse .pdf con el listado tabulado."""
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        return HttpResponse(
            'reportlab no está instalado. Ejecuta: pip install reportlab',
            status=500,
        )

    rows = build_rows(qs, fields)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(filename.upper(), styles['Title']))
    elements.append(Spacer(1, 0.4 * cm))

    table_data = [headers] + rows
    num_cols = len(headers)
    col_width = (landscape(letter)[0] - 3 * cm) / max(num_cols, 1)

    table = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 9),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME',   (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',   (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF5FB')]),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#AED6F1')),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response


def export_list_response(request, qs, filename, headers, fields):
    """
    Punto de entrada genérico para FBVs: revisa ?export=pdf|excel y
    devuelve el HttpResponse correspondiente, o None si no aplica
    (para que la vista siga su flujo normal de render).

    Uso típico en una vista de función:
        def my_list(request):
            qs = MyModel.objects.all()
            export = export_list_response(
                request, qs, 'listado_x', ['Col1', 'Col2'], ['field1', 'field2']
            )
            if export:
                return export
            return render(request, 'app/my_list.html', {'items': qs})
    """
    fmt = request.GET.get('export', '').lower()
    if fmt == 'pdf':
        return build_pdf_response(qs, filename, headers, fields)
    if fmt == 'excel':
        return build_excel_response(qs, filename, headers, fields)
    return None


# ─────────────────────────────────────────────
#  ExportMixin  —  genérico para cualquier ListView
# ─────────────────────────────────────────────
class ExportMixin:
    """
    Mixin genérico para exportar el queryset filtrado a PDF o Excel.

    Uso en una ListView:
        class MiListView(LoginRequiredMixin, ExportMixin, ListView):
            export_filename = 'mi_listado'   # sin extensión
            export_headers  = ['Col1', 'Col2', ...]
            export_fields   = ['campo1', 'campo2', ...]
            # export_fields puede contener:
            #   - nombre de atributo simple          → 'name'
            #   - lookups con doble guión bajo        → 'brand__name'
            #   - callables que reciben el objeto     → lambda obj: ...

    El mixin intercepta ?export=pdf o ?export=excel en la URL.
    Si no hay ese parámetro, sigue el flujo normal de ListView.
    """

    export_filename = 'listado'
    export_headers: list = []
    export_fields: list = []

    # ── punto de entrada ──────────────────────────────────────────────
    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('export', '').lower()
        if fmt in ('pdf', 'excel'):
            qs = self._get_export_queryset(request)
            if fmt == 'pdf':
                return build_pdf_response(qs, self.export_filename, self.export_headers, self.export_fields)
            return build_excel_response(qs, self.export_filename, self.export_headers, self.export_fields)
        return super().get(request, *args, **kwargs)

    # ── extrae queryset filtrado (reutiliza get_queryset() de la vista) ─
    def _get_export_queryset(self, request):
        # get_queryset() ya aplica los filtros definidos en la vista
        return self.get_queryset()
class GroupRequiredMixin(LoginRequiredMixin):
    """
    Mixin que permite el acceso solo a usuarios que pertenecen a ciertos grupos.
    Uso en la vista: group_required = ['Administrador', 'Vendedor']
    """
    group_required = None

    def dispatch(self, request, *args, **kwargs):
        # 1. Verificar si está logueado (heredado de LoginRequiredMixin)
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        
        # 2. Si el usuario es superusuario (admin total), pasa directo
        if request.user.is_superuser:
            return super().dispatch(request, *args, **kwargs)

        # 3. Validar si se definieron grupos requeridos
        if self.group_required:
            # Obtenemos los nombres de los grupos del usuario actual
            user_groups = request.user.groups.values_list('name', flat=True)
            
            # Verificamos si hay coincidencia entre los grupos requeridos y los del usuario
            if not any(group in user_groups for group in self.group_required):
                messages.error(request, "No tienes permisos (Rol adecuado) para acceder a esta pantalla.")
                return redirect('home') # O la url de tu página de inicio/dashboard
                
        return super().dispatch(request, *args, **kwargs)


# ─────────────────────────────────────────────
#  ModulePermissionRequiredMixin
#  Blindaje de vistas CBV a nivel de SERVIDOR usando
#  permisos dinámicos guardados en base de datos
#  (django.contrib.auth.models.Permission), en vez
#  de nombres de grupo "hardcodeados".
# ─────────────────────────────────────────────
class ModulePermissionRequiredMixin(LoginRequiredMixin, PermissionRequiredMixin):
    """
    Reemplazo recomendado de GroupRequiredMixin para las vistas nuevas.

    Se apoya en el `PermissionRequiredMixin` NATIVO de Django (que ya
    resuelve correctamente superusuarios y usa el sistema de permisos
    real de la BD), y solo le añade el comportamiento de UX que pide
    el proyecto: en vez de mostrar la página de error 403 "seca" de
    Django, redirige al Home con un mensaje de error usando
    django.contrib.messages.

    Uso en una vista (CBV):

        from shared.mixins import ModulePermissionRequiredMixin

        class PurchaseCreateView(ModulePermissionRequiredMixin, CreateView):
            model = Purchase
            ...
            permission_required = 'security.view_purchases'

    Se puede exigir más de un permiso a la vez:

        permission_required = ('security.view_purchases', 'security.view_reports')
        permission_required_all = True   # exige TODOS (por defecto True,
                                          # igual que el mixin nativo de Django)

    Si se prefiere un 403 "puro" (sin redirección, sin mensaje) en
    lugar de la redirección + mensaje, basta con setear:

        raise_exception = True

    en la vista concreta; en ese caso se delega en el comportamiento
    estándar de PermissionRequiredMixin (HttpResponseForbidden / 403).
    """

    # A dónde redirigir cuando el permiso falla (por nombre de URL).
    permission_denied_redirect = 'billing:home'
    permission_denied_message = 'No tienes permisos para acceder a este módulo.'

    # False -> UX amigable: mensaje + redirect al Home (comportamiento
    #          por defecto, tal como pide el enunciado del proyecto).
    # True  -> comportamiento "seco" nativo de Django: levanta 403.
    raise_exception = False

    def handle_no_permission(self):
        # 1. Si ni siquiera está logueado, dejamos que el flujo normal
        #    de LoginRequiredMixin lo mande a la pantalla de login.
        if not self.request.user.is_authenticated:
            return super().handle_no_permission()

        # 2. Si se pidió explícitamente el comportamiento "duro" (403),
        #    respetamos el comportamiento nativo de Django.
        if self.raise_exception:
            return super().handle_no_permission()

        # 3. Comportamiento por defecto: mensaje de error + redirect.
        messages.error(self.request, self.get_permission_denied_message())
        return redirect(self.permission_denied_redirect)


# ─────────────────────────────────────────────
#  SuccessMessageMixin
#  Automatiza los mensajes de éxito (Bootstrap 5) de
#  CreateView / UpdateView / DeleteView, detectando el
#  nombre del modelo y la acción realizada, sin tener
#  que escribir "messages.success(...)" a mano en cada vista.
# ─────────────────────────────────────────────
class SuccessMessageMixin:
    """
    Mixin genérico para vistas basadas en clases (CBV) de tipo
    CreateView, UpdateView y DeleteView.

    Al combinarlo con cualquiera de esas vistas, genera automáticamente
    un mensaje de éxito con django.contrib.messages (compatible con las
    clases 'alert alert-success' de Bootstrap 5 que ya renderiza
    templates/billing/base.html), usando el nombre "bonito" del modelo
    (Meta.verbose_name) y la acción CRUD realizada:

        - CreateView -> "El registro de <Modelo> ha sido creado exitosamente."
        - UpdateView -> "El registro de <Modelo> ha sido actualizado correctamente."
        - DeleteView -> "El registro de <Modelo> ha sido eliminado correctamente."

    Uso típico (basta con agregarlo a la herencia, ANTES de la vista
    genérica de Django, para que su form_valid()/delete() se ejecute
    primero en la cadena de super()):

        class ProductCreateView(ModulePermissionRequiredMixin, SuccessMessageMixin, CreateView):
            model = Product
            ...

        class ProductGroupDeleteView(ModulePermissionRequiredMixin, StaffRequiredMixin,
                                      SuccessMessageMixin, DeleteView):
            model = ProductGroup
            ...

    Personalización opcional (por si el mensaje automático no calza,
    por ejemplo por género gramatical -"el/la"- o por un texto más
    específico) declarando cualquiera de estos atributos en la vista:

        success_message_model_name = "Marca"      # fuerza la etiqueta a mostrar
        success_message_created = "..."           # mensaje 100% custom al crear
        success_message_updated = "..."           # mensaje 100% custom al editar
        success_message_deleted = "..."           # mensaje 100% custom al eliminar

    Si la vista ya arma su propio mensaje manualmente (como
    security.GroupCreateView o billing/purchasing FBVs), simplemente
    no se le agrega este mixin para evitar mensajes duplicados.
    """

    success_message_created = None
    success_message_updated = None
    success_message_deleted = None
    success_message_model_name = None

    def get_model_label(self):
        """Devuelve el nombre legible del modelo para el mensaje."""
        if self.success_message_model_name:
            return self.success_message_model_name

        model = getattr(self, 'model', None)
        if model is None:
            obj = getattr(self, 'object', None)
            model = obj.__class__ if obj is not None else None

        if model is not None:
            return str(model._meta.verbose_name).capitalize()
        return 'Registro'

    def form_valid(self, form):
        """
        Se dispara tanto en CreateView como en UpdateView (ambas usan
        form_valid() al guardar exitosamente). Distinguimos cuál de las
        dos es la vista actual con isinstance(), ya que 'self' es la
        instancia final de la vista concreta (ej. ProductCreateView),
        que hereda de una de esas dos clases genéricas.
        """
        response = super().form_valid(form)
        label = self.get_model_label()

        if isinstance(self, CreateView):
            message = self.success_message_created or (
                f'El registro de {label} ha sido creado exitosamente.'
            )
            messages.success(self.request, message)
        elif isinstance(self, UpdateView):
            message = self.success_message_updated or (
                f'El registro de {label} ha sido actualizado correctamente.'
            )
            messages.success(self.request, message)

        return response

    def delete(self, request, *args, **kwargs):
        """
        DeleteView NO pasa por form_valid(); su flujo de éxito es
        delete() (POST -> delete() -> redirect a success_url). Por eso
        se captura el nombre del modelo ANTES de llamar a super(), y el
        mensaje se agrega DESPUÉS de que el objeto ya fue eliminado
        realmente de la base de datos (para no felicitar al usuario si
        algo falla a mitad de camino).
        """
        label = self.get_model_label()
        response = super().delete(request, *args, **kwargs)

        message = self.success_message_deleted or (
            f'El registro de {label} ha sido eliminado correctamente.'
        )
        messages.success(request, message)
        return response